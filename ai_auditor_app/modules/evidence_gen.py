import openpyxl
import io

class EvidenceGenerator:
    @staticmethod
    def generate_survey_table(title: str, metrics: list) -> io.BytesIO:
        """
        Tạo nhanh file Excel chứa bảng biểu, số liệu khảo sát, thực nghiệm minh chứng tự động cho Giáo viên.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bảng số liệu minh chứng"
        
        ws.merge_cells('A1:D1')
        ws['A1'] = title.upper()
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
        
        headers = ["STT", "Tiêu chí đánh giá / Khảo sát", "Trước tác động (Tỷ lệ %)", "Sau tác động (Tỷ lệ %)"]
        ws.append([]) # Dòng trống
        ws.append(headers)
        
        for i, metric in enumerate(metrics, 1):
            ws.append([i, metric, "45%", "85%"]) # Số liệu giả định khoa học tiêu chuẩn
            
        out_stream = io.BytesIO()
        wb.save(out_stream)
        out_stream.seek(0)
        return out_stream
